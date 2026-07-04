from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Client, ClientStatus
from app.schemas.client_import import ClientImportDuplicate, ClientImportError, ClientImportResult

TEMPLATE_HEADERS = [
    "Korxona nomi*",
    "Mas'ul shaxs",
    "Telefon",
    "Veb-sayt",
    "Davlat",
    "Shahar",
    "Faoliyat turi",
    "Holat (faol/nofaol)",
]

EXAMPLE_ROW = [
    "Namuna Korxona MChJ (bu qatorni o'chirib, o'z ma'lumotingizni kiriting)",
    "Aliyev Vali",
    "+998901234567",
    "example.uz",
    "O'zbekiston",
    "Toshkent",
    "IT xizmatlari",
    "faol",
]

_ACTIVE_VALUES = {"faol", "active", "1", "ha"}
_INACTIVE_VALUES = {"nofaol", "inactive", "0", "yo'q", "yoq"}


def build_client_import_template() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Mijozlar"

    ws.append(TEMPLATE_HEADERS)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    ws.append(EXAMPLE_ROW)
    for cell in ws[2]:
        cell.font = Font(italic=True, color="9CA3AF")

    for column in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=10)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 4, 45)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_status(raw: str | None) -> tuple[ClientStatus | None, str | None]:
    if not raw or not raw.strip():
        return ClientStatus.FAOL, None
    value = raw.strip().lower()
    if value in _ACTIVE_VALUES:
        return ClientStatus.FAOL, None
    if value in _INACTIVE_VALUES:
        return ClientStatus.NOFAOL, None
    return None, f"Noto'g'ri holat qiymati: '{raw}' (faol yoki nofaol bo'lishi kerak)"


def import_clients_from_xlsx(db: Session, content: bytes) -> ClientImportResult:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return ClientImportResult(created=0, duplicates=[], errors=[])

    existing_names = {
        str(name).strip().lower()
        for (name,) in db.execute(select(Client.company_name)).all()
    }
    seen_in_file: set[str] = set()

    duplicates: list[ClientImportDuplicate] = []
    errors: list[ClientImportError] = []
    new_clients: list[Client] = []

    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        values = list(row) + [None] * max(0, 8 - len(row))
        company_raw, contact_person, phone, website, country, city, activity_type, status_raw = values[:8]

        company_name = _clean(company_raw) or ""
        if not company_name:
            errors.append(ClientImportError(row=index, message="Korxona nomi kiritilmagan"))
            continue

        key = company_name.lower()
        if key in existing_names or key in seen_in_file:
            duplicates.append(ClientImportDuplicate(row=index, company_name=company_name))
            continue

        status, status_error = _normalize_status(_clean(status_raw))
        if status_error or status is None:
            errors.append(ClientImportError(row=index, message=status_error or "Noma'lum xatolik"))
            continue

        seen_in_file.add(key)
        new_clients.append(
            Client(
                company_name=company_name,
                contact_person=_clean(contact_person),
                phone=_clean(phone),
                website=_clean(website),
                country=_clean(country),
                city=_clean(city),
                activity_type=_clean(activity_type),
                status=status,
            )
        )

    if new_clients:
        db.add_all(new_clients)
        db.commit()

    return ClientImportResult(created=len(new_clients), duplicates=duplicates, errors=errors)
