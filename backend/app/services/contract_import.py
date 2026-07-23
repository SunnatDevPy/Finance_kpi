import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Client, ClientStatus, Contract, ContractLineItem, Payment, ServiceType
from app.schemas.contract_import import ContractImportDuplicate, ContractImportError, ContractImportResult
from app.services.contract_status import infer_contract_workflow_status, sync_status_after_payment

TEMPLATE_HEADERS = [
    "Kompaniya / Компания*",
    "Xizmat / Услуга*",
    "Shartnoma № va sana (masalan: №1 dan 23.01.2026)",
    "Summa / Сумма*",
    "To'landi / Поступление",
    "EHF / izoh",
]

EXAMPLE_ROW = [
    "Namuna Korxona MChJ (bu qatorni o'chirib, o'z ma'lumotingizni kiriting)",
    "SMM",
    "№1 dan 23.01.2026",
    "50 000 000",
    "28 000 000",
    "17EHF",
]

# Har bir ustunni sarlavha matnidan (o'zbek/rus, turli formatlar) avtomatik aniqlash uchun
# tayanch so'zlar. Shu tufayli mijozning eski Excel fayli — ustunlar tartibi va nomlanishi
# bizning shablondan farq qilsa ham — to'g'ridan-to'g'ri yuklanaveradi.
_HEADER_ALIASES: dict[str, list[str]] = {
    "company": ["предприяти", "компани", "kompaniya", "korxona", "клиент", "mijoz"],
    "service": ["услуг", "xizmat", "xizmatlar", "hizmat"],
    "contract": ["договор", "shartnoma", "sana", "дата", "nomer"],
    "amount": ["сумма", "summa", "jami", "total"],
    "paid": ["поступлен", "tolandi", "tolang", "tolangan", "оплат", "toʻlangan"],
    "debt": ["долг", "qarz", "debt"],
    "invoice": ["эсф", "esf", "ehr", "ehf", "эхр", "эхф", "ндс", "nds", "elektron hisob", "raqami"],
}

_REQUIRED_FIELD_LABELS: dict[str, str] = {
    "company": "Kompaniya / Компания",
    "service": "Xizmat / Услуга",
    "amount": "Summa / Сумма",
}

_CLOSED_TEXT_MARKERS = (
    "готово",
    "готов",
    "tayyor",
    "yopiq",
    "yopildi",
    "done",
    "closed",
    "оплачено",
    "tolandi",
)

_DATE_RE = re.compile(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})")
_NUMBER_TOKEN_RE = re.compile(r"№?\s*([^\s№]+)")
_NUMBER_FROM_PREFIX_RE = re.compile(r"(?:№|#|Nº|nº)?\s*(\d+)", re.UNICODE)
_NON_NUMERIC_RE = re.compile(r"[^\d.,-]")
_HEADER_STRIP_RE = re.compile(r"[^\w\sʻʼ]", re.UNICODE)
_HEADER_SCAN_LIMIT = 20


def build_contract_import_template() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Shartnomalar"

    ws.append(TEMPLATE_HEADERS)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    ws.append(EXAMPLE_ROW)
    for cell in ws[2]:
        cell.font = Font(italic=True, color="9CA3AF")

    for column in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=10)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 4, 55)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_header(value: Any) -> str:
    return _HEADER_STRIP_RE.sub("", str(value or "").strip().lower())


def _detect_columns(header_row: tuple[Any, ...]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        text = _normalize_header(raw)
        if not text:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if field in columns:
                continue
            if any(alias in text for alias in aliases):
                columns[field] = idx
                break
    return columns


def _score_header_row(header_row: tuple[Any, ...]) -> int:
    columns = _detect_columns(header_row)
    required = {"company", "service", "amount"}
    score = len(columns)
    if required.issubset(columns):
        score += 10
    return score


def _find_header_row(ws) -> tuple[int, tuple[Any, ...]]:
    best_row = 1
    best_score = -1
    best_data: tuple[Any, ...] = ()
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=_HEADER_SCAN_LIMIT, values_only=True),
        start=1,
    ):
        score = _score_header_row(row or ())
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_data = row or ()
    return best_row, best_data


def _looks_closed(raw: Any) -> bool:
    if raw is None:
        return False
    text = _normalize_header(raw)
    return any(marker in text for marker in _CLOSED_TEXT_MARKERS)


def _parse_amount(raw: Any) -> Decimal | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float, Decimal)):
        try:
            value = Decimal(str(raw))
        except InvalidOperation:
            return None
        return value if value != 0 else None

    text = str(raw).strip()
    if not text:
        return None
    cleaned = _NON_NUMERIC_RE.sub("", text).replace(",", "")
    if cleaned.count(".") > 1:
        cleaned = cleaned.replace(".", "")
    if not cleaned or cleaned in {"-", "."}:
        return None
    try:
        value = Decimal(cleaned)
    except InvalidOperation:
        return None
    return value if value != 0 else None


def _parse_contract_number_and_date(raw: Any, fallback: date) -> tuple[str | None, date]:
    if raw is None:
        return None, fallback
    if isinstance(raw, datetime):
        return None, raw.date()
    if isinstance(raw, date):
        return None, raw

    text = str(raw).strip()
    if not text:
        return None, fallback

    parsed_date = fallback
    remainder = text
    date_match = _DATE_RE.search(text)
    if date_match:
        day, month, year = (int(group) for group in date_match.groups())
        if year < 100:
            year += 2000
        try:
            parsed_date = date(year, month, day)
        except ValueError:
            parsed_date = fallback
        remainder = text[: date_match.start()] + text[date_match.end() :]

    for word in ("dan", "от", "from", "-", ":"):
        remainder = remainder.replace(word, " ")
    remainder = re.sub(r"\s+", " ", remainder).strip()

    contract_number: str | None = None
    prefix_match = _NUMBER_FROM_PREFIX_RE.search(remainder)
    if prefix_match:
        contract_number = prefix_match.group(1)
    else:
        number_match = _NUMBER_TOKEN_RE.search(remainder)
        if number_match:
            contract_number = number_match.group(1).strip()
            digit_match = re.search(r"\d+", contract_number)
            if digit_match:
                contract_number = digit_match.group(0)

    return (contract_number or None), parsed_date


def import_contracts_from_xlsx(db: Session, content: bytes) -> ContractImportResult:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return ContractImportResult(
            created_contracts=0, created_clients=0, created_service_types=0, duplicates=[], errors=[]
        )

    header_row_index, header_row = _find_header_row(ws)
    columns = _detect_columns(header_row)
    missing = [label for field, label in _REQUIRED_FIELD_LABELS.items() if field not in columns]
    if missing:
        raise ValueError(
            "Quyidagi ustunlar aniqlanmadi: "
            + ", ".join(missing)
            + ". Sarlavha qatorida shu nomlar (yoki shablondagidek) bo'lishi kerak."
        )

    def cell(row: tuple[Any, ...], field: str) -> Any:
        idx = columns.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    today = date.today()

    clients_by_name: dict[str, Client] = {
        client.company_name.strip().lower(): client
        for client in db.execute(select(Client)).scalars().all()
    }
    service_types_by_name: dict[str, ServiceType] = {
        service.name.strip().lower(): service
        for service in db.execute(select(ServiceType)).scalars().all()
    }
    existing_contract_keys: set[tuple[int, str]] = {
        (contract.client_id, contract.contract_number.strip().lower())
        for contract in db.execute(select(Contract)).scalars().all()
        if contract.contract_number
    }

    created_clients = 0
    created_service_types = 0
    duplicates: list[ContractImportDuplicate] = []
    errors: list[ContractImportError] = []
    new_contracts: list[Contract] = []

    for index, row in enumerate(
        ws.iter_rows(min_row=header_row_index + 1, values_only=True),
        start=header_row_index + 1,
    ):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        company_name = _clean(cell(row, "company"))
        service_name = _clean(cell(row, "service"))
        amount = _parse_amount(cell(row, "amount"))

        if not company_name:
            errors.append(ContractImportError(row=index, message="Kompaniya nomi kiritilmagan"))
            continue
        if not service_name:
            errors.append(ContractImportError(row=index, message="Xizmat turi kiritilmagan"))
            continue
        if amount is None or amount <= 0:
            errors.append(ContractImportError(row=index, message="Summa noto'g'ri yoki kiritilmagan"))
            continue

        contract_number, contract_date = _parse_contract_number_and_date(cell(row, "contract"), today)
        invoice_number = _clean(cell(row, "invoice"))

        # Сумма == Поступление bo'lsa — shartnoma to'liq yopilgan hisoblanadi. Agar "To'landi"
        # ustuni bo'sh qoldirilgan bo'lsa-yu, "Долг" ustunida "готово"/"tayyor" kabi holat
        # yozilgan bo'lsa, buni ham to'liq to'lov sifatida hisobga olamiz (rang emas, matn/summa
        # taqqoslashga tayaniladi).
        paid_amount = _parse_amount(cell(row, "paid"))
        if paid_amount is None:
            paid_amount = amount if _looks_closed(cell(row, "debt")) else Decimal("0")
        elif paid_amount >= amount:
            paid_amount = amount

        client_key = company_name.lower()
        client = clients_by_name.get(client_key)
        if client is None:
            client = Client(company_name=company_name, status=ClientStatus.FAOL)
            db.add(client)
            db.flush()
            clients_by_name[client_key] = client
            created_clients += 1

        if contract_number:
            dedup_key = (client.id, contract_number.strip().lower())
            if dedup_key in existing_contract_keys:
                duplicates.append(
                    ContractImportDuplicate(
                        row=index, company_name=company_name, contract_number=contract_number
                    )
                )
                continue
            existing_contract_keys.add(dedup_key)

        service_key = service_name.lower()
        service_type = service_types_by_name.get(service_key)
        if service_type is None:
            service_type = ServiceType(name=service_name, is_active=True)
            db.add(service_type)
            db.flush()
            service_types_by_name[service_key] = service_type
            created_service_types += 1

        contract = Contract(
            client=client,
            start_date=contract_date,
            end_date=contract_date,
            contract_number=contract_number,
            invoice_number=invoice_number,
            notes="Excel import orqali qo'shilgan",
            line_items=[ContractLineItem(service_type=service_type, price=amount)],
        )
        contract.status = infer_contract_workflow_status(contract)
        if paid_amount > 0:
            contract.payments.append(
                Payment(amount=paid_amount, paid_at=contract_date, note="Excel import orqali")
            )
        sync_status_after_payment(contract)

        db.add(contract)
        new_contracts.append(contract)

    if new_contracts:
        db.commit()

    return ContractImportResult(
        created_contracts=len(new_contracts),
        created_clients=created_clients,
        created_service_types=created_service_types,
        duplicates=duplicates,
        errors=errors,
    )
