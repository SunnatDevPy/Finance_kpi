from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.services.export_data import EXPORT_TITLES
from app.services.pdf_fonts import ensure_unicode_fonts


def build_xlsx(title: str, headers: list[str], rows: list[list[str]]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(row)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_client_card_xlsx(
    company_name: str,
    sheets: list[tuple[str, list[str], list[list[str]]]],
) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    for index, (title, headers, rows) in enumerate(sheets):
        ws = wb.create_sheet(title=title[:31], index=index)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append(row)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 48)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_pdf(title: str, headers: list[str], rows: list[list[str]]) -> BytesIO:
    font_regular, font_bold = ensure_unicode_fonts()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), title=title)
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontName = font_bold
    elements = [
        Paragraph(title, title_style),
        Spacer(1, 12),
    ]

    table_data = [headers, *rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), font_bold),
                ("FONTNAME", (0, 1), (-1, -1), font_regular),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_resource_file(
    resource: str,
    file_format: str,
    headers: list[str],
    rows: list[list[str]],
) -> tuple[BytesIO, str, str]:
    title = EXPORT_TITLES[resource]
    ext = "xlsx" if file_format == "xlsx" else "pdf"
    filename = f"{resource}_{title.lower().replace(' ', '_')}.{ext}"

    if file_format == "xlsx":
        buffer = build_xlsx(title, headers, rows)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        buffer = build_pdf(title, headers, rows)
        media_type = "application/pdf"

    return buffer, filename, media_type
