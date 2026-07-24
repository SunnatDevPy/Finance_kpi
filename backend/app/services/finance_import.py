import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Literal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.models import Expense, ExpenseCategory, Income, IncomeCategory
from app.schemas.finance_import import FinanceImportError, FinanceImportResult

TEMPLATE_HEADERS = [
    "Sana* / Дата",
    "Turi* (Kirim/Chiqim) / Тип",
    "Summa* / Сумма",
    "Kategoriya / Категория",
    "Izoh / Комментарий",
]

EXAMPLE_ROWS = [
    [
        "23.01.2026",
        "Kirim",
        "5 000 000",
        "Sotuv",
        "Mijozdan naqd tushum (bu qatorni o'chirib, o'z ma'lumotingizni kiriting)",
    ],
    [
        "24.01.2026",
        "Chiqim",
        "3 000 000",
        "Ijara",
        "Ofis ijarasi — namuna qator",
    ],
]

# Ustunlarni sarlavha matnidan (o'zbek/rus, turli formatlar) avtomatik aniqlash uchun
# tayanch so'zlar — shu tufayli eski Excel fayl ustunlari shablon bilan bir xil
# tartibda/nomlanishda bo'lmasa ham to'g'ridan-to'g'ri yuklanaveradi.
_HEADER_ALIASES: dict[str, list[str]] = {
    "date": ["sana", "дата", "date"],
    "type": ["turi", "тип", "type"],
    "title": ["nomi", "tavsif", "название", "наименование", "title"],
    "amount": ["сумма", "summa", "amount"],
    "category": ["kategoriya", "категория", "category"],
    "note": ["izoh", "коммент", "eslatma", "note", "comment"],
}

_REQUIRED_FIELD_LABELS: dict[str, str] = {
    "date": "Sana / Дата",
    "amount": "Summa / Сумма",
}

_INCOME_MARKERS = ("kirim", "приход", "income", "tushum", "keldi", "доход")
_EXPENSE_MARKERS = ("chiqim", "расход", "expense", "xarajat", "chiqdi")

_SKIP_ROW_MARKERS = (
    "bu qatorni o'chirib",
    "удалите эту строку",
    "namuna qator",
    "пример строки",
)

_INCOME_DEFAULT_TITLES: dict[IncomeCategory, str] = {
    IncomeCategory.SALE: "Sotuv",
    IncomeCategory.SERVICE: "Xizmat",
    IncomeCategory.INVESTMENT: "Investitsiya",
    IncomeCategory.LOAN: "Kredit",
    IncomeCategory.GRANT: "Grant",
    IncomeCategory.REFUND: "Qaytarma",
    IncomeCategory.OTHER: "Boshqa kirim",
}

_EXPENSE_DEFAULT_TITLES: dict[ExpenseCategory, str] = {
    ExpenseCategory.SALARY: "Ish haqi",
    ExpenseCategory.RENT: "Ijara",
    ExpenseCategory.MARKETING: "Marketing",
    ExpenseCategory.UTILITIES: "Kommunal",
    ExpenseCategory.TRANSPORT: "Transport",
    ExpenseCategory.OFFICE: "Ofis",
    ExpenseCategory.TAX: "Soliq",
    ExpenseCategory.BANK_FEE: "Bank to'lovi",
    ExpenseCategory.OTHER: "Boshqa xarajat",
}

_EXPENSE_CATEGORY_ALIASES: dict[ExpenseCategory, list[str]] = {
    ExpenseCategory.SALARY: ["ish haqi", "oylik", "zarplata", "зарплат", "salary"],
    ExpenseCategory.RENT: ["ijara", "аренд", "rent"],
    ExpenseCategory.MARKETING: ["marketing", "реклам", "smm"],
    ExpenseCategory.UTILITIES: ["kommunal", "коммунал", "utilit"],
    ExpenseCategory.TRANSPORT: ["transport", "транспорт", "yoqilg'i", "benzin"],
    ExpenseCategory.OFFICE: ["ofis", "офис", "office", "kanselyariya"],
    ExpenseCategory.TAX: ["soliq", "налог", "tax"],
    ExpenseCategory.BANK_FEE: ["bank", "банк"],
}

_INCOME_CATEGORY_ALIASES: dict[IncomeCategory, list[str]] = {
    IncomeCategory.SALE: ["sotuv", "продаж", "sale"],
    IncomeCategory.SERVICE: ["xizmat", "услуг", "service"],
    IncomeCategory.INVESTMENT: ["investitsiya", "инвестиц", "investment"],
    IncomeCategory.LOAN: ["kredit", "qarz", "кредит", "заем", "loan"],
    IncomeCategory.GRANT: ["grant", "грант"],
    IncomeCategory.REFUND: ["qaytar", "возврат", "refund"],
}

_DATE_RE = re.compile(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})")
_NON_NUMERIC_RE = re.compile(r"[^\d.,-]")
_HEADER_STRIP_RE = re.compile(r"[^\w\sʻʼ]", re.UNICODE)


def build_finance_import_template() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Moliya tarixi"

    ws.append(TEMPLATE_HEADERS)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for example in EXAMPLE_ROWS:
        ws.append(example)
    for row in ws.iter_rows(min_row=2, max_row=1 + len(EXAMPLE_ROWS)):
        for cell in row:
            cell.font = Font(italic=True, color="9CA3AF")

    for column in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=10)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 4, 55)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize(value: Any) -> str:
    return _HEADER_STRIP_RE.sub("", str(value or "").strip().lower())


def _detect_columns(header_row: tuple[Any, ...]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        text = _normalize(raw)
        if not text:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if field in columns:
                continue
            if any(alias in text for alias in aliases):
                columns[field] = idx
                break
    return columns


def _parse_amount(raw: Any) -> Decimal | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float, Decimal)):
        try:
            value = Decimal(str(raw))
        except InvalidOperation:
            return None
        return value if value != 0 else None

    text = str(raw).strip()
    if not text:
        return None
    cleaned = _NON_NUMERIC_RE.sub("", text).replace(",", "")
    if cleaned.count(".") > 1:
        cleaned = cleaned.replace(".", "")
    if not cleaned or cleaned in {"-", "."}:
        return None
    try:
        value = Decimal(cleaned)
    except InvalidOperation:
        return None
    return value if value != 0 else None


def _parse_date(raw: Any, fallback: date) -> date:
    if raw is None:
        return fallback
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw

    text = str(raw).strip()
    if not text:
        return fallback
    match = _DATE_RE.search(text)
    if not match:
        return fallback
    day, month, year = (int(group) for group in match.groups())
    if year < 100:
        year += 2000
    try:
        return date(year, month, day)
    except ValueError:
        return fallback


def _detect_type(raw: Any, amount: Decimal | None) -> Literal["income", "expense"] | None:
    if raw is not None:
        text = _normalize(raw)
        if any(marker in text for marker in _EXPENSE_MARKERS):
            return "expense"
        if any(marker in text for marker in _INCOME_MARKERS):
            return "income"
    if amount is not None:
        return "expense" if amount < 0 else "income"
    return None


def _resolve_expense_category(raw: Any) -> ExpenseCategory:
    text = _normalize(raw)
    if text:
        for category, aliases in _EXPENSE_CATEGORY_ALIASES.items():
            if any(alias in text for alias in aliases):
                return category
    return ExpenseCategory.OTHER


def _resolve_income_category(raw: Any) -> IncomeCategory:
    text = _normalize(raw)
    if text:
        for category, aliases in _INCOME_CATEGORY_ALIASES.items():
            if any(alias in text for alias in aliases):
                return category
    return IncomeCategory.OTHER


def _is_skip_row(row: tuple[Any, ...]) -> bool:
    for raw in row:
        if raw is None:
            continue
        text = str(raw).strip().lower()
        if any(marker in text for marker in _SKIP_ROW_MARKERS):
            return True
    return False


def _resolve_entry_content(
    *,
    note_raw: str | None,
    legacy_title: str | None,
    entry_type: Literal["income", "expense"],
    income_category: IncomeCategory,
    expense_category: ExpenseCategory,
) -> tuple[str, str | None]:
    text = (note_raw or legacy_title or "").strip()
    if text:
        return text[:255], text
    if entry_type == "income":
        return _INCOME_DEFAULT_TITLES.get(income_category, "Kirim")[:255], None
    return _EXPENSE_DEFAULT_TITLES.get(expense_category, "Xarajat")[:255], None


def import_finance_from_xlsx(db: Session, content: bytes) -> FinanceImportResult:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return FinanceImportResult(created_income=0, created_expense=0, errors=[])

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    columns = _detect_columns(header_row or ())
    missing = [label for field, label in _REQUIRED_FIELD_LABELS.items() if field not in columns]
    if missing:
        raise ValueError(
            "Quyidagi ustunlar aniqlanmadi: "
            + ", ".join(missing)
            + ". Sarlavha qatorida shu nomlar (yoki shablondagidek) bo'lishi kerak."
        )

    def cell(row: tuple[Any, ...], field: str) -> Any:
        idx = columns.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    today = date.today()
    new_incomes: list[Income] = []
    new_expenses: list[Expense] = []
    errors: list[FinanceImportError] = []

    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        if _is_skip_row(row):
            continue

        amount = _parse_amount(cell(row, "amount"))
        if amount is None or amount == 0:
            errors.append(FinanceImportError(row=index, message="Summa noto'g'ri yoki kiritilmagan"))
            continue

        entry_type = _detect_type(cell(row, "type"), amount)
        if entry_type is None:
            errors.append(
                FinanceImportError(
                    row=index,
                    message="Kirim/chiqim turi aniqlanmadi (\"Turi\" ustuniga Kirim yoki Chiqim deb yozing)",
                )
            )
            continue

        entry_date = _parse_date(cell(row, "date"), today)
        income_category = _resolve_income_category(cell(row, "category"))
        expense_category = _resolve_expense_category(cell(row, "category"))
        title, note = _resolve_entry_content(
            note_raw=_clean(cell(row, "note")),
            legacy_title=_clean(cell(row, "title")),
            entry_type=entry_type,
            income_category=income_category,
            expense_category=expense_category,
        )
        magnitude = abs(amount)

        if entry_type == "income":
            new_incomes.append(
                Income(
                    category=income_category,
                    title=title,
                    amount=magnitude,
                    income_date=entry_date,
                    note=note,
                )
            )
        else:
            new_expenses.append(
                Expense(
                    category=expense_category,
                    title=title,
                    amount=magnitude,
                    expense_date=entry_date,
                    note=note,
                )
            )

    if new_incomes:
        db.add_all(new_incomes)
    if new_expenses:
        db.add_all(new_expenses)
    if new_incomes or new_expenses:
        db.commit()

    return FinanceImportResult(
        created_income=len(new_incomes),
        created_expense=len(new_expenses),
        errors=errors,
    )
