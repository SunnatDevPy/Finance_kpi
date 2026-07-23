from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

REAL_2019_WORKBOOK = Path(r"c:\Users\user\Downloads\Telegram Desktop\2019.xlsx")


def _build_xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["Kompaniya", "Xizmat", "Shartnoma", "Summa", "To'landi", "ЭСФ"])
    for row in rows:
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _upload(client, auth_headers, content: bytes):
    return client.post(
        "/api/v1/contracts/import",
        headers=auth_headers,
        files={
            "file": (
                "shartnomalar.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def test_download_contract_import_template(client, auth_headers):
    response = client.get("/api/v1/contracts/import-template", headers=auth_headers)
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats")


def test_import_contracts_creates_client_contract_and_payment(client, auth_headers):
    content = _build_xlsx(
        [
            ["Chust Aroma MChJ", "Брендбук", "№1 dan 23.01.2026", "50 000 000", "", ""],
            ["Gurlan Global Teks", "Сайт", "№5 dan 17.02.2026", "28 000 000", "28 000 000", "17ЭСФ"],
        ]
    )
    response = _upload(client, auth_headers, content)
    assert response.status_code == 200
    data = response.json()
    assert data["created_contracts"] == 2
    assert data["created_clients"] == 2
    assert data["created_service_types"] == 2
    assert data["duplicates"] == []
    assert data["errors"] == []

    clients_resp = client.get("/api/v1/clients", headers=auth_headers, params={"limit": 50}).json()
    client_row = next(c for c in clients_resp["items"] if c["company_name"] == "Gurlan Global Teks")

    contracts_resp = client.get(
        "/api/v1/contracts", headers=auth_headers, params={"client_id": client_row["id"]}
    ).json()
    assert contracts_resp["total"] == 1
    contract = contracts_resp["items"][0]
    assert contract["contract_number"] == "5"
    assert contract["invoice_number"] == "17ЭСФ"
    assert contract["start_date"] == "2026-02-17"
    assert float(contract["total_amount"]) == 28_000_000.0
    assert float(contract["paid_amount"]) == 28_000_000.0
    assert float(contract["debt_amount"]) == 0.0


def test_import_contracts_reuses_existing_client_and_service(
    client, auth_headers, sample_client, sample_service_type
):
    content = _build_xlsx(
        [
            [
                sample_client.company_name,
                sample_service_type.name,
                "№9 dan 01.03.2026",
                "10 000 000",
                "4 000 000",
                "",
            ],
        ]
    )
    response = _upload(client, auth_headers, content)
    assert response.status_code == 200
    data = response.json()
    assert data["created_contracts"] == 1
    assert data["created_clients"] == 0
    assert data["created_service_types"] == 0


def test_import_contracts_detects_duplicate_by_contract_number(client, auth_headers):
    content = _build_xlsx(
        [
            ["Zelal Tekstil", "Video", "№12 dan 17.04.2026", "56 000 000", "", ""],
        ]
    )
    first = _upload(client, auth_headers, content)
    assert first.json()["created_contracts"] == 1

    second = _upload(client, auth_headers, content)
    data = second.json()
    assert data["created_contracts"] == 0
    assert len(data["duplicates"]) == 1
    assert data["duplicates"][0]["contract_number"] == "12"


def test_import_contracts_reports_row_errors(client, auth_headers):
    content = _build_xlsx(
        [
            ["", "SMM", "№1 dan 01.01.2026", "1 000 000", "", ""],
            ["Kompaniya Y", "", "№2 dan 01.01.2026", "1 000 000", "", ""],
            ["Kompaniya Z", "SMM", "№3 dan 01.01.2026", "noto'g'ri summa", "", ""],
        ]
    )
    response = _upload(client, auth_headers, content)
    data = response.json()
    assert response.status_code == 200
    assert data["created_contracts"] == 0
    assert len(data["errors"]) == 3


def test_import_contracts_detects_real_world_header_layout(client, auth_headers):
    """Mijozning haqiqiy Excel fayli — ustunlar boshqa tartibda va rus tilida nomlangan,
    oldida qo'shimcha \u2116 ustuni bor, "\u0414\u043e\u043b\u0433" ustuni esa e'tiborga olinmasligi kerak."""
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "№",
            "Наименование предприятия",
            "Название услуги",
            "Номер и дата договора",
            "Сумма",
            "Поступление",
            "Долг",
        ]
    )
    ws.append([1, "Bakan Tex", "Каталог", "№1 от 06.01.2020", 4000000, 4000000, "готово"])
    ws.append([2, "Buyuk Turon Fayz", "Логотип", "№5 от 03.02.2020", 2000000, 1000000, 1000000])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = _upload(client, auth_headers, buffer.read())
    assert response.status_code == 200
    data = response.json()
    assert data["created_contracts"] == 2
    assert data["created_clients"] == 2
    assert data["errors"] == []

    contracts = client.get("/api/v1/contracts", headers=auth_headers, params={"limit": 50}).json()["items"]
    bakan = next(c for c in contracts if c["contract_number"] == "1")
    assert float(bakan["debt_amount"]) == 0.0
    turon = next(c for c in contracts if c["contract_number"] == "5")
    assert float(turon["debt_amount"]) == 1_000_000.0


def test_import_contracts_treats_closed_status_text_as_full_payment(client, auth_headers):
    """Agar 'Поступление' bo'sh qoldirilgan bo'lsa-yu, 'Долг' ustunida 'готово' kabi holat
    yozilgan bo'lsa — shartnoma to'liq to'langan deb hisoblanadi (rang emas, matnga qarab)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Компания", "Услуга", "Договор", "Сумма", "Поступление", "Долг"])
    ws.append(["Fully Closed MChJ", "SMM", "№9 от 01.01.2026", 3_000_000, None, "готово"])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = _upload(client, auth_headers, buffer.read())
    assert response.status_code == 200
    data = response.json()
    assert data["created_contracts"] == 1

    contracts = client.get("/api/v1/contracts", headers=auth_headers, params={"limit": 50}).json()["items"]
    contract = next(c for c in contracts if c["contract_number"] == "9")
    assert float(contract["paid_amount"]) == 3_000_000.0
    assert float(contract["debt_amount"]) == 0.0


def test_import_contracts_supports_2019_export_layout(client, auth_headers):
    """WTMA eksporti: 1-qator bo'sh, 2-qatorda B–H ustunlari (Mijoz, Sana, Jami, ...)."""
    wb = Workbook()
    ws = wb.active
    ws.append([None] * 8)
    ws.append(
        [
            None,
            "Mijoz",
            "Sana",
            "Jami",
            "To'langan",
            "Qarz",
            "Xizmatlar",
            "EHR raqami",
        ]
    )
    ws.append([None, "Surkhon Teks", "№1  17.12.2019", 4_000_000, 4_000_000, None, "Katalog", None])
    ws.append([None, "Azalium", "№2  27.12.2019", 2_225_000, 2_225_000, None, "Sayt", "EHF-002"])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = _upload(client, auth_headers, buffer.read())
    assert response.status_code == 200
    data = response.json()
    assert data["created_contracts"] == 2
    assert data["created_clients"] == 2
    assert data["errors"] == []

    contracts = client.get("/api/v1/contracts", headers=auth_headers, params={"limit": 50}).json()["items"]
    first = next(c for c in contracts if c["contract_number"] == "1")
    assert first["start_date"] == "2019-12-17"
    assert first["status"] == "tugadi"
    assert float(first["paid_amount"]) == 4_000_000.0
    second = next(c for c in contracts if c["contract_number"] == "2")
    assert second["invoice_number"] == "EHF-002"
    assert second["status"] == "tugadi"


def test_import_contracts_parses_sana_with_contract_number_prefix(client, auth_headers):
    content = _build_xlsx(
        [
            ["Mijoz A", "Sayt", "№15  15.08.2026", "1 000 000", "500 000", "", ""],
            ["Mijoz B", "SMM", "№3 dan 11.01.2026", "2 000 000", "2 000 000", "", ""],
        ]
    )
    response = _upload(client, auth_headers, content)
    assert response.status_code == 200
    data = response.json()
    assert data["created_contracts"] == 2

    contracts = client.get("/api/v1/contracts", headers=auth_headers, params={"limit": 50}).json()["items"]
    partial = next(c for c in contracts if c["contract_number"] == "15")
    assert partial["start_date"] == "2026-08-15"
    assert partial["status"] == "yangi"
    assert float(partial["debt_amount"]) == 500_000.0
    paid = next(c for c in contracts if c["contract_number"] == "3")
    assert paid["start_date"] == "2026-01-11"
    assert paid["status"] == "tugadi"


def test_import_real_2019_workbook_when_available(client, auth_headers):
    if not REAL_2019_WORKBOOK.is_file():
        return

    response = _upload(client, auth_headers, REAL_2019_WORKBOOK.read_bytes())
    assert response.status_code == 200
    data = response.json()
    assert data["created_contracts"] == 3
    assert data["created_clients"] == 3
    assert data["errors"] == []

    contracts = client.get("/api/v1/contracts", headers=auth_headers, params={"limit": 50}).json()["items"]
    surkhon = next(c for c in contracts if c["contract_number"] == "1")
    assert surkhon["start_date"] == "2019-12-17"
    assert surkhon["status"] == "tugadi"
    assert float(surkhon["debt_amount"]) == 0.0


def test_import_contracts_rejects_unrecognized_headers(client, auth_headers):
    wb = Workbook()
    ws = wb.active
    ws.append(["Column A", "Column B", "Column C"])
    ws.append(["Random", "Data", "Row"])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = _upload(client, auth_headers, buffer.read())
    assert response.status_code == 400
    assert "ustunlar aniqlanmadi" in response.json()["detail"].lower()


def test_import_contracts_rejects_non_xlsx(client, auth_headers):
    response = client.post(
        "/api/v1/contracts/import",
        headers=auth_headers,
        files={"file": ("shartnomalar.csv", b"a,b,c", "text/csv")},
    )
    assert response.status_code == 400
