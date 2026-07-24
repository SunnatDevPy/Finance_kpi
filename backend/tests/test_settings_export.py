from io import BytesIO

from openpyxl import load_workbook


def test_export_all_returns_xlsx_workbook(client, auth_headers, sample_contract):
    client.post(
        "/api/v1/payments",
        headers=auth_headers,
        json={
            "contract_id": sample_contract.id,
            "amount": "500000.00",
            "paid_at": "2026-03-15",
        },
    )
    client.post(
        "/api/v1/incomes",
        headers=auth_headers,
        json={
            "category": "sale",
            "title": "Qo'lda tushum",
            "amount": "200000.00",
            "income_date": "2026-03-20",
        },
    )
    client.post(
        "/api/v1/expenses",
        headers=auth_headers,
        json={
            "category": "rent",
            "title": "Ijara",
            "amount": "100000.00",
            "expense_date": "2026-03-21",
        },
    )

    response = client.get("/api/v1/settings/export-all", headers=auth_headers)
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = load_workbook(BytesIO(response.content))
    assert "Mijozlar" in wb.sheetnames
    assert "Kontraktlar" in wb.sheetnames
    assert "To'lovlar" in wb.sheetnames
    assert "Kirimlar" in wb.sheetnames
    assert "Xarajatlar" in wb.sheetnames
    assert "Xodimlar" in wb.sheetnames

    clients_sheet = wb["Mijozlar"]
    assert clients_sheet.cell(row=2, column=1).value == "Acme LLC"

    employees_sheet = wb["Xodimlar"]
    assert employees_sheet.cell(row=2, column=2).value == "admin"


def test_export_all_with_date_range(client, auth_headers, sample_contract):
    client.post(
        "/api/v1/payments",
        headers=auth_headers,
        json={
            "contract_id": sample_contract.id,
            "amount": "500000.00",
            "paid_at": "2026-03-15",
        },
    )

    response = client.get(
        "/api/v1/settings/export-all",
        headers=auth_headers,
        params={"date_from": "2026-01-01", "date_to": "2026-12-31"},
    )
    assert response.status_code == 200


def test_export_all_requires_admin(client, db_session):
    from app.models import User, UserRole
    from app.services.auth import hash_password

    user = User(
        username="menejer1",
        full_name="Menejer",
        password_hash=hash_password("password123"),
        role=UserRole.MENEJER,
        is_active=True,
    )
    db_session.add(user)
    db_session.commit()

    login = client.post(
        "/api/v1/auth/login",
        json={"username": "menejer1", "password": "password123"},
    )
    token = login.json()["access_token"]

    response = client.get(
        "/api/v1/settings/export-all",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 403
